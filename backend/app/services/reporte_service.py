"""Reportes mensuales y estadisticas.

Todo se calcula agregando los `datos_snapshot` inmutables de cada
`GeneracionPedido`, por lo que no requiere tablas ni columnas nuevas: un pedido
ya guarda, al confirmarse, su `resumen_global` (una fila por combinacion
ingrediente-localidad-proveedor con cantidad y costo) y su `costo_total`.

Un pedido se imputa al mes de su `semana_inicio` (el lunes de la semana para
REGULAR; la fecha del sabado/evento para PATIO/EVENTO). Las semanas que cruzan
el cambio de mes se cuentan enteras en el mes de su lunes: es una regla simple y
predecible, acorde a como se opera el SAE.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.models.pedido_model import GeneracionPedido
from app.services.pedido_service import (
    _commercial_quantity_label,
    _commercial_unit_label,
    _dec,
    _money,
    _qty,
    _safe_filename,
    _titulo,
    list_pedidos,
)

_MESES_ES = [
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

_TIPO_LABELS = {
    "REGULAR": "Semanal",
    "PATIO": "Patios",
    "EVENTO": "Eventos",
}

_TIPOS_VALIDOS = tuple(_TIPO_LABELS.keys())


def _tipo_label(tipo: str | None) -> str:
    if not tipo:
        return ""
    return _TIPO_LABELS.get(tipo, tipo.title())


def _month_label(anio: int, mes: int) -> str:
    if 1 <= mes <= 12:
        return f"{_MESES_ES[mes]} {anio}"
    return f"{mes}/{anio}"


def _percent(part: Decimal, total: Decimal) -> float:
    if total <= 0:
        return 0.0
    return float((part / total * Decimal("100")).quantize(Decimal("0.1")))


def _validate_tipo(tipo: str | None) -> str | None:
    if tipo is None:
        return None
    normalized = tipo.strip().upper()
    if normalized not in _TIPOS_VALIDOS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tipo de pedido invalido (REGULAR, PATIO o EVENTO)",
        )
    return normalized


def _load_pedidos(db: Session, tipo: str | None) -> list[GeneracionPedido]:
    return list_pedidos(db, tipo=_validate_tipo(tipo))


def _in_month(pedido: GeneracionPedido, anio: int, mes: int) -> bool:
    fecha = pedido.semana_inicio
    return fecha is not None and fecha.year == anio and fecha.month == mes


def _pedido_costo(snapshot: dict) -> Decimal:
    return _dec(snapshot.get("costo_total", "0") or "0")


# ---------------------------------------------------------------------------
# Motor de agregacion (comun a reporte mensual y estadisticas)
# ---------------------------------------------------------------------------

class _Aggregate:
    """Acumula magnitudes en Decimal a partir de un conjunto de pedidos.

    Se recorre cada pedido una sola vez: el desglose por ingrediente/localidad/
    proveedor sale de `resumen_global` (que ya trae costo por combinacion) y el
    costo por escuela de la seccion `escuelas` del snapshot.
    """

    def __init__(self) -> None:
        self.resumen: dict[tuple, dict] = {}
        self.por_localidad: dict[object, dict] = {}
        self.por_proveedor: dict[object, dict] = {}
        self.por_ingrediente: dict[object, dict] = {}
        self.por_escuela: dict[object, dict] = {}
        self.por_tipo: dict[str, dict] = {}
        self.por_mes: dict[tuple[int, int], dict] = {}
        self.escuelas_con_costo: set = set()
        self.num_pedidos = 0
        self.costo_total = Decimal("0")

    def add(self, pedido: GeneracionPedido) -> None:
        snapshot = pedido.datos_snapshot or {}
        self.num_pedidos += 1
        costo_pedido = _pedido_costo(snapshot)
        self.costo_total += costo_pedido

        tipo = snapshot.get("tipo") or pedido.tipo or "REGULAR"
        tipo_entry = self.por_tipo.setdefault(
            tipo, {"tipo": tipo, "tipo_label": _tipo_label(tipo), "costo": Decimal("0"), "num": 0}
        )
        tipo_entry["costo"] += costo_pedido
        tipo_entry["num"] += 1

        fecha: date | None = pedido.semana_inicio
        if fecha is not None:
            mes_key = (fecha.year, fecha.month)
            mes_entry = self.por_mes.setdefault(
                mes_key, {"anio": fecha.year, "mes": fecha.month, "costo": Decimal("0"), "num": 0}
            )
            mes_entry["costo"] += costo_pedido
            mes_entry["num"] += 1

        for row in snapshot.get("resumen_global", []):
            self._add_resumen_row(row)

        for escuela in snapshot.get("escuelas", []):
            self._add_escuela(escuela)

    def _add_resumen_row(self, row: dict) -> None:
        cantidad = _dec(row.get("cantidad_total", "0") or "0")
        costo = _dec(row.get("costo_total", "0") or "0")
        contenido_valor = row.get("cantidad_contenido_total")
        contenido = _dec(contenido_valor) if contenido_valor not in (None, "") else None

        ingrediente_id = row.get("ingrediente_id")
        localidad_id = row.get("localidad_id")
        proveedor_id = row.get("proveedor_id")

        key = (ingrediente_id, localidad_id, proveedor_id)
        entry = self.resumen.get(key)
        if entry is None:
            entry = {
                "ingrediente_id": ingrediente_id,
                "ingrediente_nombre": row.get("ingrediente_nombre", ""),
                "unidad": row.get("unidad", ""),
                "contenido_por_unidad": row.get("contenido_por_unidad"),
                "unidad_contenido": row.get("unidad_contenido"),
                "localidad_id": localidad_id,
                "localidad_nombre": row.get("localidad_nombre", ""),
                "proveedor_id": proveedor_id,
                "proveedor_nombre": row.get("proveedor_nombre", ""),
                "cantidad_total": Decimal("0"),
                "cantidad_contenido_total": Decimal("0"),
                "tiene_contenido": row.get("contenido_por_unidad") is not None,
                "costo_total": Decimal("0"),
            }
            self.resumen[key] = entry
        entry["cantidad_total"] += cantidad
        if contenido is not None:
            entry["cantidad_contenido_total"] += contenido
        entry["costo_total"] += costo

        loc = self.por_localidad.setdefault(
            localidad_id,
            {"localidad_id": localidad_id, "localidad_nombre": row.get("localidad_nombre", ""), "costo": Decimal("0")},
        )
        loc["costo"] += costo

        prov = self.por_proveedor.setdefault(
            proveedor_id,
            {
                "proveedor_id": proveedor_id,
                "proveedor_nombre": row.get("proveedor_nombre", ""),
                "costo": Decimal("0"),
                "localidades": set(),
            },
        )
        prov["costo"] += costo
        if row.get("localidad_nombre"):
            prov["localidades"].add(row.get("localidad_nombre"))

        ing = self.por_ingrediente.setdefault(
            ingrediente_id,
            {"ingrediente_id": ingrediente_id, "ingrediente_nombre": row.get("ingrediente_nombre", ""), "costo": Decimal("0")},
        )
        ing["costo"] += costo

    def _add_escuela(self, escuela: dict) -> None:
        costo = Decimal("0")
        for item in escuela.get("ingredientes", []):
            if "proveedor_id" not in item:
                continue
            costo += _dec(item.get("costo_total", "0") or "0")

        escuela_id = escuela.get("escuela_id")
        entry = self.por_escuela.setdefault(
            escuela_id,
            {
                "escuela_id": escuela_id,
                "codigo": escuela.get("codigo", ""),
                "nombre": escuela.get("nombre", ""),
                "localidad_nombre": escuela.get("localidad_nombre", ""),
                "costo": Decimal("0"),
            },
        )
        entry["costo"] += costo
        if costo > 0:
            self.escuelas_con_costo.add(escuela_id)


def _aggregate(pedidos: list[GeneracionPedido]) -> _Aggregate:
    agg = _Aggregate()
    for pedido in pedidos:
        agg.add(pedido)
    return agg


# ---------------------------------------------------------------------------
# Serializacion a estructuras de respuesta
# ---------------------------------------------------------------------------

def _resumen_rows(agg: _Aggregate) -> list[dict]:
    rows = []
    for entry in agg.resumen.values():
        cantidad = entry["cantidad_total"]
        costo = entry["costo_total"]
        precio_promedio = costo / cantidad if cantidad > 0 else Decimal("0")
        rows.append(
            {
                "ingrediente_id": entry["ingrediente_id"],
                "ingrediente_nombre": entry["ingrediente_nombre"],
                "unidad": entry["unidad"],
                "contenido_por_unidad": entry["contenido_por_unidad"],
                "unidad_contenido": entry["unidad_contenido"],
                "localidad_id": entry["localidad_id"],
                "localidad_nombre": entry["localidad_nombre"],
                "proveedor_id": entry["proveedor_id"],
                "proveedor_nombre": entry["proveedor_nombre"],
                "cantidad_total": _qty(cantidad),
                "cantidad_contenido_total": (
                    _qty(entry["cantidad_contenido_total"]) if entry["tiene_contenido"] else None
                ),
                "precio_promedio": _money(precio_promedio),
                "costo_total": _money(costo),
            }
        )
    rows.sort(key=lambda r: (r["ingrediente_nombre"].lower(), r["localidad_nombre"].lower()))
    return rows


def _localidad_rows(agg: _Aggregate) -> list[dict]:
    total = agg.costo_total
    rows = [
        {
            "localidad_id": entry["localidad_id"],
            "localidad_nombre": entry["localidad_nombre"],
            "costo_total": _money(entry["costo"]),
            "porcentaje": _percent(entry["costo"], total),
        }
        for entry in agg.por_localidad.values()
    ]
    rows.sort(key=lambda r: _dec(r["costo_total"]), reverse=True)
    return rows


def _proveedor_rows(agg: _Aggregate) -> list[dict]:
    total = agg.costo_total
    rows = [
        {
            "proveedor_id": entry["proveedor_id"],
            "proveedor_nombre": entry["proveedor_nombre"],
            "localidades": ", ".join(sorted(entry["localidades"])),
            "costo_total": _money(entry["costo"]),
            "porcentaje": _percent(entry["costo"], total),
        }
        for entry in agg.por_proveedor.values()
    ]
    rows.sort(key=lambda r: _dec(r["costo_total"]), reverse=True)
    return rows


def _escuela_rows(agg: _Aggregate) -> list[dict]:
    total = agg.costo_total
    rows = [
        {
            "escuela_id": entry["escuela_id"],
            "codigo": entry["codigo"],
            "nombre": entry["nombre"],
            "localidad_nombre": entry["localidad_nombre"],
            "costo_total": _money(entry["costo"]),
            "porcentaje": _percent(entry["costo"], total),
        }
        for entry in agg.por_escuela.values()
        if entry["costo"] > 0
    ]
    rows.sort(key=lambda r: _dec(r["costo_total"]), reverse=True)
    return rows


def _ingrediente_rows(agg: _Aggregate, limit: int | None = None) -> list[dict]:
    total = agg.costo_total
    rows = [
        {
            "ingrediente_id": entry["ingrediente_id"],
            "ingrediente_nombre": entry["ingrediente_nombre"],
            "costo_total": _money(entry["costo"]),
            "porcentaje": _percent(entry["costo"], total),
        }
        for entry in agg.por_ingrediente.values()
    ]
    rows.sort(key=lambda r: _dec(r["costo_total"]), reverse=True)
    if limit is not None:
        rows = rows[:limit]
    return rows


def _tipo_rows(agg: _Aggregate) -> list[dict]:
    total = agg.costo_total
    rows = [
        {
            "tipo": entry["tipo"],
            "tipo_label": entry["tipo_label"],
            "num_pedidos": entry["num"],
            "costo_total": _money(entry["costo"]),
            "porcentaje": _percent(entry["costo"], total),
        }
        for entry in agg.por_tipo.values()
    ]
    rows.sort(key=lambda r: _dec(r["costo_total"]), reverse=True)
    return rows


def _pedido_incluido(pedido: GeneracionPedido) -> dict:
    snapshot = pedido.datos_snapshot or {}
    tipo = snapshot.get("tipo") or pedido.tipo or "REGULAR"
    return {
        "id": pedido.id,
        "fecha": pedido.semana_inicio.isoformat() if pedido.semana_inicio else "",
        "tipo": tipo,
        "tipo_label": _tipo_label(tipo),
        "detalle": _titulo(snapshot),
        "costo_total": _money(_pedido_costo(snapshot)),
    }


# ---------------------------------------------------------------------------
# API publica del servicio
# ---------------------------------------------------------------------------

def list_meses(db: Session, tipo: str | None = None) -> list[dict]:
    """Meses con al menos un pedido, del mas reciente al mas antiguo."""
    pedidos = _load_pedidos(db, tipo)
    buckets: dict[tuple[int, int], dict] = {}
    for pedido in pedidos:
        fecha = pedido.semana_inicio
        if fecha is None:
            continue
        key = (fecha.year, fecha.month)
        entry = buckets.setdefault(
            key,
            {
                "anio": fecha.year,
                "mes": fecha.month,
                "etiqueta": _month_label(fecha.year, fecha.month),
                "num_pedidos": 0,
                "costo": Decimal("0"),
            },
        )
        entry["num_pedidos"] += 1
        entry["costo"] += _pedido_costo(pedido.datos_snapshot or {})

    meses = sorted(buckets.values(), key=lambda m: (m["anio"], m["mes"]), reverse=True)
    return [
        {
            "anio": m["anio"],
            "mes": m["mes"],
            "etiqueta": m["etiqueta"],
            "num_pedidos": m["num_pedidos"],
            "costo_total": _money(m["costo"]),
        }
        for m in meses
    ]


def build_reporte_mensual(db: Session, anio: int, mes: int, tipo: str | None = None) -> dict:
    if not (1 <= mes <= 12):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Mes invalido (1-12)")
    tipo = _validate_tipo(tipo)
    pedidos = [p for p in _load_pedidos(db, tipo) if _in_month(p, anio, mes)]
    agg = _aggregate(pedidos)

    pedidos_incluidos = [_pedido_incluido(p) for p in pedidos]
    pedidos_incluidos.sort(key=lambda p: p["fecha"], reverse=True)

    return {
        "anio": anio,
        "mes": mes,
        "etiqueta": _month_label(anio, mes),
        "tipo": tipo,
        "num_pedidos": agg.num_pedidos,
        "costo_total": _money(agg.costo_total),
        "resumen": _resumen_rows(agg),
        "por_proveedor": _proveedor_rows(agg),
        "por_localidad": _localidad_rows(agg),
        "por_escuela": _escuela_rows(agg),
        "por_tipo": _tipo_rows(agg),
        "pedidos": pedidos_incluidos,
    }


def build_estadisticas(db: Session, anio: int | None = None, tipo: str | None = None) -> dict:
    tipo = _validate_tipo(tipo)
    todos = _load_pedidos(db, tipo)
    anios_disponibles = sorted(
        {p.semana_inicio.year for p in todos if p.semana_inicio is not None}, reverse=True
    )

    pedidos = [p for p in todos if anio is None or (p.semana_inicio and p.semana_inicio.year == anio)]
    agg = _aggregate(pedidos)

    tendencia = [
        {
            "anio": entry["anio"],
            "mes": entry["mes"],
            "etiqueta": _month_label(entry["anio"], entry["mes"]),
            "costo_total": _money(entry["costo"]),
            "num_pedidos": entry["num"],
        }
        for entry in sorted(agg.por_mes.values(), key=lambda m: (m["anio"], m["mes"]))
    ]

    mes_pico = max(agg.por_mes.values(), key=lambda m: m["costo"], default=None)
    costo_promedio = agg.costo_total / agg.num_pedidos if agg.num_pedidos else Decimal("0")

    totales = {
        "costo_total": _money(agg.costo_total),
        "num_pedidos": agg.num_pedidos,
        "num_escuelas": len(agg.escuelas_con_costo),
        "num_proveedores": len(agg.por_proveedor),
        "num_localidades": len(agg.por_localidad),
        "costo_promedio_pedido": _money(costo_promedio),
        "mes_pico_etiqueta": _month_label(mes_pico["anio"], mes_pico["mes"]) if mes_pico else None,
        "mes_pico_costo": _money(mes_pico["costo"]) if mes_pico else None,
    }

    return {
        "anio": anio,
        "tipo": tipo,
        "anios": anios_disponibles,
        "totales": totales,
        "tendencia": tendencia,
        "por_localidad": _localidad_rows(agg),
        "por_proveedor": _proveedor_rows(agg),
        "top_ingredientes": _ingrediente_rows(agg, limit=15),
        "por_tipo": _tipo_rows(agg),
    }


# ---------------------------------------------------------------------------
# Exportacion del reporte mensual (Excel / PDF)
# ---------------------------------------------------------------------------

def _reporte_filename(anio: int, mes: int, tipo: str | None, extension: str) -> str:
    parts = ["reporte_mensual", f"{anio}-{mes:02d}"]
    if tipo:
        parts.append(tipo.lower())
    return _safe_filename("_".join(parts)) + f".{extension}"


def export_reporte_mensual_excel(db: Session, anio: int, mes: int, tipo: str | None = None) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    reporte = build_reporte_mensual(db, anio, mes, tipo)
    workbook = Workbook()

    resumen = workbook.active
    resumen.title = "Resumen mensual"
    resumen.append(["REPORTE MENSUAL SAE"])
    resumen.append(["Mes", reporte["etiqueta"]])
    resumen.append(["Tipo", _tipo_label(reporte["tipo"]) or "Todos"])
    resumen.append(["Pedidos incluidos", reporte["num_pedidos"]])
    resumen.append(["Costo total del mes", reporte["costo_total"]])
    resumen.append([])
    resumen.append(
        ["Ingrediente", "Unidad", "Localidad", "Cantidad total", "Proveedor", "Precio prom.", "Costo total"]
    )
    for cell in resumen[7]:
        cell.font = Font(bold=True)
    for row in reporte["resumen"]:
        resumen.append(
            [
                row["ingrediente_nombre"],
                _commercial_unit_label(row),
                row["localidad_nombre"],
                _commercial_quantity_label(row),
                row["proveedor_nombre"],
                row["precio_promedio"],
                row["costo_total"],
            ]
        )
    if reporte["resumen"]:
        resumen.append(["TOTAL", "", "", "", "", "", reporte["costo_total"]])
        for cell in resumen[resumen.max_row]:
            cell.font = Font(bold=True)

    prov_sheet = workbook.create_sheet("Por proveedor")
    prov_sheet.append(["Proveedor", "Localidades", "Costo total", "% del mes"])
    for cell in prov_sheet[1]:
        cell.font = Font(bold=True)
    for row in reporte["por_proveedor"]:
        prov_sheet.append([row["proveedor_nombre"], row["localidades"], row["costo_total"], row["porcentaje"]])

    loc_sheet = workbook.create_sheet("Por localidad")
    loc_sheet.append(["Localidad", "Costo total", "% del mes"])
    for cell in loc_sheet[1]:
        cell.font = Font(bold=True)
    for row in reporte["por_localidad"]:
        loc_sheet.append([row["localidad_nombre"], row["costo_total"], row["porcentaje"]])

    esc_sheet = workbook.create_sheet("Por escuela")
    esc_sheet.append(["Codigo", "Escuela", "Localidad", "Costo total", "% del mes"])
    for cell in esc_sheet[1]:
        cell.font = Font(bold=True)
    for row in reporte["por_escuela"]:
        esc_sheet.append(
            [row["codigo"], row["nombre"], row["localidad_nombre"], row["costo_total"], row["porcentaje"]]
        )

    pedidos_sheet = workbook.create_sheet("Pedidos")
    pedidos_sheet.append(["Fecha", "Tipo", "Detalle", "Costo total"])
    for cell in pedidos_sheet[1]:
        cell.font = Font(bold=True)
    for row in reporte["pedidos"]:
        pedidos_sheet.append([row["fecha"], row["tipo_label"], row["detalle"], row["costo_total"]])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 45)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_reporte_mensual_pdf(db: Session, anio: int, mes: int, tipo: str | None = None) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    reporte = build_reporte_mensual(db, anio, mes, tipo)
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("REPORTE MENSUAL SAE", styles["Title"]),
        Paragraph(f"Mes: {reporte['etiqueta']}", styles["Normal"]),
        Paragraph(f"Tipo: {_tipo_label(reporte['tipo']) or 'Todos'}", styles["Normal"]),
        Paragraph(f"Pedidos incluidos: {reporte['num_pedidos']}", styles["Normal"]),
        Paragraph(f"Costo total del mes: {reporte['costo_total']}", styles["Normal"]),
        Spacer(1, 12),
    ]

    header_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]

    resumen_data = [["Ingrediente", "Unidad", "Localidad", "Cant. total", "Proveedor", "Precio prom.", "Costo total"]]
    for row in reporte["resumen"]:
        resumen_data.append(
            [
                row["ingrediente_nombre"],
                _commercial_unit_label(row),
                row["localidad_nombre"],
                _commercial_quantity_label(row),
                row["proveedor_nombre"],
                row["precio_promedio"],
                row["costo_total"],
            ]
        )
    if len(resumen_data) > 1:
        resumen_data.append(["TOTAL", "", "", "", "", "", reporte["costo_total"]])
    else:
        resumen_data.append(["Sin datos", "", "", "", "", "", ""])
    resumen_style = list(header_style)
    if reporte["resumen"]:
        resumen_style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
    resumen_table = Table(resumen_data, repeatRows=1)
    resumen_table.setStyle(TableStyle(resumen_style))
    story.append(Paragraph("Resumen consolidado", styles["Heading2"]))
    story.append(resumen_table)

    if reporte["por_proveedor"]:
        story.append(Spacer(1, 14))
        story.append(Paragraph("Costo por proveedor", styles["Heading2"]))
        prov_data = [["Proveedor", "Localidades", "Costo total", "% del mes"]]
        for row in reporte["por_proveedor"]:
            prov_data.append([row["proveedor_nombre"], row["localidades"], row["costo_total"], f"{row['porcentaje']}%"])
        prov_table = Table(prov_data, repeatRows=1)
        prov_table.setStyle(TableStyle(header_style))
        story.append(prov_table)

    if reporte["por_localidad"]:
        story.append(Spacer(1, 14))
        story.append(Paragraph("Costo por localidad", styles["Heading2"]))
        loc_data = [["Localidad", "Costo total", "% del mes"]]
        for row in reporte["por_localidad"]:
            loc_data.append([row["localidad_nombre"], row["costo_total"], f"{row['porcentaje']}%"])
        loc_table = Table(loc_data, repeatRows=1)
        loc_table.setStyle(TableStyle(header_style))
        story.append(loc_table)

    if reporte["pedidos"]:
        story.append(Spacer(1, 14))
        story.append(Paragraph("Pedidos incluidos", styles["Heading2"]))
        pedido_data = [["Fecha", "Tipo", "Detalle", "Costo total"]]
        for row in reporte["pedidos"]:
            pedido_data.append([row["fecha"], row["tipo_label"], row["detalle"], row["costo_total"]])
        pedido_table = Table(pedido_data, repeatRows=1)
        pedido_table.setStyle(TableStyle(header_style))
        story.append(pedido_table)

    doc.build(story)
    output.seek(0)
    return output
