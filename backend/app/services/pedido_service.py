import json
import re
import zipfile
from io import BytesIO
from datetime import date
from decimal import Decimal, ROUND_CEILING

from fastapi import HTTPException, status
from sqlalchemy.orm import Session, selectinload

from app.controllers.pedido_controller import ConfirmPedidoRequest, PreviewPedidoRequest
from app.models.asignacion_proveedor_model import AsignacionProveedor
from app.models.pedido_model import GeneracionPedido
from app.models.receta_model import Receta, RecetaIngrediente
from app.models.school_model import School
from app.models.stock_previo_model import StockPrevio
from app.models.temporada_model import DiaMenu, OpcionMenu
from app.models.user_model import User, UserRole


def _dec(value: Decimal | int | str) -> Decimal:
    return Decimal(str(value))


def _money(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01")))


def _qty(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01")))


def _ceil_decimal(value: Decimal) -> Decimal:
    return value.to_integral_value(rounding=ROUND_CEILING)


def _calculate_order_quantity(
    cantidad_neta: Decimal,
    unidad_medida: str,
    contenido_por_unidad: Decimal | None = None,
    unidad_contenido: str | None = None,
) -> dict[str, Decimal | str | None]:
    if unidad_medida.strip().lower() != "unidades":
        # Se redondea siempre hacia arriba al entero mas cercano: el proveedor
        # entrega cantidades enteras de kg/litros/gramos/docenas (por escuela).
        return {
            "cantidad_final": _ceil_decimal(cantidad_neta),
            "unidad_final": unidad_medida,
            "unidad_calculo": unidad_medida,
            "contenido_por_unidad": None,
            "unidad_contenido": None,
            "cantidad_contenido_final": None,
        }

    if contenido_por_unidad is None or _dec(contenido_por_unidad) <= 0 or not unidad_contenido:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El ingrediente por unidades debe tener contenido y unidad de contenido validos",
        )

    contenido = _dec(contenido_por_unidad)
    cantidad_final = _ceil_decimal(cantidad_neta / contenido)
    return {
        "cantidad_final": cantidad_final,
        "unidad_final": unidad_medida,
        "unidad_calculo": unidad_contenido,
        "contenido_por_unidad": contenido,
        "unidad_contenido": unidad_contenido,
        "cantidad_contenido_final": cantidad_final * contenido,
    }


def _validate_dias(dias_habiles: list[int]) -> list[int]:
    if not dias_habiles:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Debe seleccionarse al menos un dia habil")
    if any(dia < 1 or dia > 5 for dia in dias_habiles):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Dias habiles invalidos")
    return sorted(set(dias_habiles))


def _load_opcion(db: Session, opcion_menu_id: int) -> OpcionMenu:
    opcion = db.query(OpcionMenu).filter(OpcionMenu.id == opcion_menu_id).first()
    if not opcion:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Opcion de menu no encontrada")
    return opcion


def _load_menu_rows(
    db: Session,
    opcion_menu_id: int,
    dias_habiles: list[int],
) -> list[DiaMenu]:
    rows = (
        db.query(DiaMenu)
        .options(
            selectinload(DiaMenu.receta)
            .selectinload(Receta.ingredientes)
            .joinedload(RecetaIngrediente.ingrediente)
        )
        .filter(
            DiaMenu.opcion_menu_id == opcion_menu_id,
            DiaMenu.dia_semana.in_(dias_habiles),
        )
        .order_by(DiaMenu.dia_semana, DiaMenu.tipo_comida_id)
        .all()
    )
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La opcion de menu no tiene recetas para los dias seleccionados",
        )
    return rows


def _active_provider(
    db: Session,
    ingrediente_id: int,
    localidad_id: int,
    reference_date: date,
) -> AsignacionProveedor | None:
    return (
        db.query(AsignacionProveedor)
        .filter(
            AsignacionProveedor.ingrediente_id == ingrediente_id,
            AsignacionProveedor.localidad_id == localidad_id,
            AsignacionProveedor.fecha_desde <= reference_date,
            (
                (AsignacionProveedor.fecha_hasta == None)
                | (AsignacionProveedor.fecha_hasta > reference_date)
            ),
        )
        .order_by(AsignacionProveedor.fecha_desde.desc(), AsignacionProveedor.id.desc())
        .first()
    )


def _stock_map(
    db: Session,
    overrides,
) -> dict[tuple[int, int], Decimal]:
    rows = db.query(StockPrevio).all()
    stock = {(row.escuela_id, row.ingrediente_id): _dec(row.cantidad) for row in rows}
    for item in overrides:
        stock[(item.escuela_id, item.ingrediente_id)] = _dec(item.cantidad)
    return stock


def _format_option(opcion: OpcionMenu) -> dict:
    temporada = opcion.temporada
    return {
        "id": opcion.id,
        "numero_opcion": opcion.numero_opcion,
        "descripcion": opcion.descripcion,
        "temporada": {
            "id": temporada.id if temporada else None,
            "nombre": temporada.nombre.value if temporada else None,
            "anio": temporada.anio if temporada else None,
        },
    }


def build_preview_snapshot(
    db: Session,
    data: PreviewPedidoRequest | ConfirmPedidoRequest,
) -> dict:
    dias_habiles = _validate_dias(data.dias_habiles)
    opcion = _load_opcion(db, data.opcion_menu_id)
    menu_rows = _load_menu_rows(db, data.opcion_menu_id, dias_habiles)
    schools = (
        db.query(School)
        .options(selectinload(School.tipos_comida))
        .filter(School.active == True)
        .order_by(School.name)
        .all()
    )
    if not schools:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay escuelas activas")

    stock_by_key = _stock_map(db, data.stock_overrides)
    provider_reference_date = data.semana_inicio
    advertencias: list[dict] = []
    escuelas_snapshot: list[dict] = []
    provider_groups: dict[str, dict] = {}
    global_rows: dict[str, dict] = {}

    for school in schools:
        base_by_ingredient: dict[int, dict] = {}
        offered_tipo_ids = {tipo.id for tipo in school.tipos_comida}

        for row in menu_rows:
            if row.tipo_comida_id not in offered_tipo_ids:
                continue

            if not row.receta or not row.receta.activo:
                continue

            for recipe_item in row.receta.ingredientes:
                ingredient = recipe_item.ingrediente
                if ingredient is None or not ingredient.activo:
                    continue

                entry = base_by_ingredient.setdefault(
                    ingredient.id,
                    {
                        "ingrediente": ingredient,
                        "cantidad_base": Decimal("0"),
                    },
                )
                entry["cantidad_base"] += _dec(recipe_item.cantidad_por_porcion) * _dec(school.matriculation)

        school_items = []
        for ingredient_id, entry in sorted(
            base_by_ingredient.items(),
            key=lambda item: item[1]["ingrediente"].nombre,
        ):
            ingredient = entry["ingrediente"]
            cantidad_base = entry["cantidad_base"]
            cantidad_corregida = cantidad_base * _dec(ingredient.indice_correccion)
            stock = stock_by_key.get((school.id, ingredient_id), Decimal("0"))
            cantidad_neta = max(Decimal("0"), cantidad_corregida - stock)

            order_quantity = _calculate_order_quantity(
                cantidad_neta,
                ingredient.unidad_medida,
                ingredient.contenido_por_unidad,
                ingredient.unidad_contenido,
            )
            cantidad_final = order_quantity["cantidad_final"]
            unidad_final = order_quantity["unidad_final"]
            unidad_calculo = order_quantity["unidad_calculo"]

            assignment = _active_provider(
                db,
                ingredient_id,
                school.locality_id,
                provider_reference_date,
            )

            item_snapshot = {
                "ingrediente_id": ingredient.id,
                "ingrediente_nombre": ingredient.nombre,
                "unidad_calculo": unidad_calculo,
                "unidad_final": unidad_final,
                "cantidad_base": _qty(cantidad_base),
                "cantidad_corregida": _qty(cantidad_corregida),
                "stock_descontado": _qty(stock),
                "cantidad_neta": _qty(cantidad_neta),
                "cantidad_final": _qty(cantidad_final),
            }
            if order_quantity["contenido_por_unidad"] is not None:
                item_snapshot.update(
                    {
                        "contenido_por_unidad": _qty(order_quantity["contenido_por_unidad"]),
                        "unidad_contenido": order_quantity["unidad_contenido"],
                        "cantidad_contenido_final": _qty(
                            order_quantity["cantidad_contenido_final"]
                        ),
                    }
                )

            if assignment is None:
                advertencias.append(
                    {
                        "tipo": "SIN_PROVEEDOR",
                        "escuela_id": school.id,
                        "escuela_nombre": school.name,
                        "localidad_id": school.locality_id,
                        "localidad_nombre": school.locality.nombre if school.locality else "",
                        "ingrediente_id": ingredient.id,
                        "ingrediente_nombre": ingredient.nombre,
                    }
                )
            else:
                provider = assignment.proveedor
                locality = assignment.localidad
                precio = _dec(assignment.precio_unitario)
                costo = cantidad_final * precio
                item_snapshot.update(
                    {
                        "proveedor_id": assignment.proveedor_id,
                        "proveedor_nombre": provider.nombre if provider else "",
                        "localidad_id": assignment.localidad_id,
                        "localidad_nombre": locality.nombre if locality else "",
                        "precio_unitario": _money(precio),
                        "costo_total": _money(costo),
                    }
                )

                group_key = f"{assignment.proveedor_id}:{assignment.localidad_id}"
                group = provider_groups.setdefault(
                    group_key,
                    {
                        "proveedor_id": assignment.proveedor_id,
                        "proveedor_nombre": provider.nombre if provider else "",
                        "localidad_id": assignment.localidad_id,
                        "localidad_nombre": locality.nombre if locality else "",
                        "ingredientes": {},
                    },
                )
                ingredient_group = group["ingredientes"].setdefault(
                    str(ingredient.id),
                    {
                        "ingrediente_id": ingredient.id,
                        "ingrediente_nombre": ingredient.nombre,
                        "unidad": unidad_final,
                        "contenido_por_unidad": item_snapshot.get("contenido_por_unidad"),
                        "unidad_contenido": item_snapshot.get("unidad_contenido"),
                        "precio_unitario": _money(precio),
                        "cantidad_total": Decimal("0"),
                        "cantidad_contenido_total": Decimal("0"),
                        "costo_total": Decimal("0"),
                        "escuelas": [],
                    },
                )
                ingredient_group["cantidad_total"] += cantidad_final
                if order_quantity["cantidad_contenido_final"] is not None:
                    ingredient_group["cantidad_contenido_total"] += order_quantity[
                        "cantidad_contenido_final"
                    ]
                ingredient_group["costo_total"] += costo
                ingredient_group["escuelas"].append(
                    {
                        "escuela_id": school.id,
                        "escuela_codigo": school.code,
                        "escuela_nombre": school.name,
                        "cantidad": _qty(cantidad_final),
                        "cantidad_contenido": (
                            _qty(order_quantity["cantidad_contenido_final"])
                            if order_quantity["cantidad_contenido_final"] is not None
                            else None
                        ),
                    }
                )

                global_key = f"{ingredient.id}:{assignment.localidad_id}:{assignment.proveedor_id}"
                global_row = global_rows.setdefault(
                    global_key,
                    {
                        "ingrediente_id": ingredient.id,
                        "ingrediente_nombre": ingredient.nombre,
                        "unidad": unidad_final,
                        "contenido_por_unidad": item_snapshot.get("contenido_por_unidad"),
                        "unidad_contenido": item_snapshot.get("unidad_contenido"),
                        "localidad_id": assignment.localidad_id,
                        "localidad_nombre": locality.nombre if locality else "",
                        "proveedor_id": assignment.proveedor_id,
                        "proveedor_nombre": provider.nombre if provider else "",
                        "precio_unitario": _money(precio),
                        "cantidad_total": Decimal("0"),
                        "cantidad_contenido_total": Decimal("0"),
                        "costo_total": Decimal("0"),
                    },
                )
                global_row["cantidad_total"] += cantidad_final
                if order_quantity["cantidad_contenido_final"] is not None:
                    global_row["cantidad_contenido_total"] += order_quantity[
                        "cantidad_contenido_final"
                    ]
                global_row["costo_total"] += costo

            school_items.append(item_snapshot)

        escuelas_snapshot.append(
            {
                "escuela_id": school.id,
                "codigo": school.code,
                "nombre": school.name,
                "localidad_id": school.locality_id,
                "localidad_nombre": school.locality.nombre if school.locality else "",
                "matricula": school.matriculation,
                "ingredientes": school_items,
            }
        )

    proveedores = []
    for group in provider_groups.values():
        ingredientes = []
        for item in group["ingredientes"].values():
            ingredientes.append(
                {
                    **{
                        k: v
                        for k, v in item.items()
                        if k not in ("cantidad_total", "cantidad_contenido_total", "costo_total")
                    },
                    "cantidad_total": _qty(item["cantidad_total"]),
                    "cantidad_contenido_total": (
                        _qty(item["cantidad_contenido_total"])
                        if item["contenido_por_unidad"] is not None
                        else None
                    ),
                    "costo_total": _money(item["costo_total"]),
                }
            )
        proveedores.append({**{k: v for k, v in group.items() if k != "ingredientes"}, "ingredientes": ingredientes})

    resumen_global = []
    costo_total = Decimal("0")
    for row in global_rows.values():
        costo_total += row["costo_total"]
        resumen_global.append(
            {
                **{
                    k: v
                    for k, v in row.items()
                    if k not in ("cantidad_total", "cantidad_contenido_total", "costo_total")
                },
                "cantidad_total": _qty(row["cantidad_total"]),
                "cantidad_contenido_total": (
                    _qty(row["cantidad_contenido_total"])
                    if row["contenido_por_unidad"] is not None
                    else None
                ),
                "costo_total": _money(row["costo_total"]),
            }
        )

    return {
        "semana_inicio": data.semana_inicio.isoformat(),
        "dias_habiles": dias_habiles,
        "opcion_menu": _format_option(opcion),
        "fecha_referencia_proveedores": provider_reference_date.isoformat(),
        "escuelas": escuelas_snapshot,
        "proveedores": proveedores,
        "resumen_global": resumen_global,
        "advertencias": advertencias,
        "costo_total": _money(costo_total),
    }


def preview_pedido(db: Session, data: PreviewPedidoRequest) -> dict:
    return build_preview_snapshot(db, data)


def get_existing_pedido_by_week(db: Session, semana_inicio: date) -> GeneracionPedido | None:
    return (
        db.query(GeneracionPedido)
        .filter(GeneracionPedido.semana_inicio == semana_inicio)
        .order_by(GeneracionPedido.id.desc())
        .first()
    )


def confirm_pedido(db: Session, data: ConfirmPedidoRequest, user: User) -> GeneracionPedido:
    existing = get_existing_pedido_by_week(db, data.semana_inicio)
    if existing is not None:
        return existing

    snapshot = build_preview_snapshot(db, data)
    pedido = GeneracionPedido(
        semana_inicio=data.semana_inicio,
        opcion_menu_id=data.opcion_menu_id,
        dias_habiles=json.dumps(snapshot["dias_habiles"]),
        generado_por_id=user.id,
        notas=data.notas,
        datos_snapshot=snapshot,
    )
    db.add(pedido)

    school_ids = [school["escuela_id"] for school in snapshot["escuelas"]]
    if school_ids:
        db.query(StockPrevio).filter(StockPrevio.escuela_id.in_(school_ids)).update(
            {StockPrevio.cantidad: Decimal("0")},
            synchronize_session=False,
        )

    db.commit()
    db.refresh(pedido)
    return pedido


def list_pedidos(db: Session) -> list[GeneracionPedido]:
    return (
        db.query(GeneracionPedido)
        .order_by(GeneracionPedido.semana_inicio.desc(), GeneracionPedido.id.desc())
        .all()
    )


def _snapshot_has_school(snapshot: dict, school_id: int | None) -> bool:
    if school_id is None:
        return False
    return any(school.get("escuela_id") == school_id for school in snapshot.get("escuelas", []))


def list_pedidos_for_user(db: Session, user: User) -> list[GeneracionPedido]:
    pedidos = list_pedidos(db)
    if user.role in (UserRole.admin, UserRole.gestor):
        return pedidos
    if user.role == UserRole.escuela:
        return [pedido for pedido in pedidos if _snapshot_has_school(pedido.datos_snapshot, user.school_id)]
    return []


def get_pedido_by_id(db: Session, pedido_id: int) -> GeneracionPedido:
    pedido = db.query(GeneracionPedido).filter(GeneracionPedido.id == pedido_id).first()
    if not pedido:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Pedido no encontrado")
    return pedido


def get_pedido_for_user(db: Session, pedido_id: int, user: User) -> GeneracionPedido:
    pedido = get_pedido_by_id(db, pedido_id)
    if user.role in (UserRole.admin, UserRole.gestor):
        return pedido
    if user.role == UserRole.escuela and _snapshot_has_school(pedido.datos_snapshot, user.school_id):
        return pedido
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso denegado")


def school_snapshot_for_user(snapshot: dict, user: User) -> dict:
    if user.role in (UserRole.admin, UserRole.gestor):
        return snapshot
    if user.role != UserRole.escuela or user.school_id is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso denegado")

    schools = [
        school for school in snapshot.get("escuelas", [])
        if school.get("escuela_id") == user.school_id
    ]
    if not schools:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso denegado")

    school_ids = {school["escuela_id"] for school in schools}
    filtered = {**snapshot, "escuelas": schools}

    proveedores = []
    for provider in snapshot.get("proveedores", []):
        ingredients = []
        for ingredient in provider.get("ingredientes", []):
            school_rows = [
                row for row in ingredient.get("escuelas", [])
                if row.get("escuela_id") in school_ids
            ]
            if school_rows:
                ingredients.append({**ingredient, "escuelas": school_rows})
        if ingredients:
            proveedores.append({**provider, "ingredientes": ingredients})

    filtered["proveedores"] = proveedores
    filtered["advertencias"] = [
        warning for warning in snapshot.get("advertencias", [])
        if warning.get("escuela_id") in school_ids
    ]

    resumen = []
    total = Decimal("0")
    for school in schools:
        for item in school.get("ingredientes", []):
            if "proveedor_id" not in item:
                continue
            cost = _dec(item.get("costo_total", "0"))
            total += cost
            resumen.append({
                "ingrediente_id": item.get("ingrediente_id"),
                "ingrediente_nombre": item.get("ingrediente_nombre", ""),
                "unidad": item.get("unidad_final", ""),
                "localidad_id": item.get("localidad_id"),
                "localidad_nombre": item.get("localidad_nombre", ""),
                "proveedor_id": item.get("proveedor_id"),
                "proveedor_nombre": item.get("proveedor_nombre", ""),
                "precio_unitario": item.get("precio_unitario", ""),
                "cantidad_total": item.get("cantidad_final", ""),
                "contenido_por_unidad": item.get("contenido_por_unidad"),
                "unidad_contenido": item.get("unidad_contenido"),
                "cantidad_contenido_total": item.get("cantidad_contenido_final"),
                "costo_total": item.get("costo_total", ""),
            })
    filtered["resumen_global"] = resumen
    filtered["costo_total"] = _money(total)
    return filtered


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("_") or "documento"


def _commercial_unit_label(item: dict) -> str:
    unit = item.get("unidad", item.get("unidad_final", ""))
    content = item.get("contenido_por_unidad")
    content_unit = item.get("unidad_contenido")
    if content and content_unit:
        return f"{unit} ({content} {content_unit} c/u)"
    return unit


def _commercial_quantity_label(
    item: dict,
    quantity_key: str = "cantidad_total",
    content_key: str = "cantidad_contenido_total",
) -> str:
    quantity = item.get(quantity_key, "")
    content = item.get(content_key)
    content_unit = item.get("unidad_contenido")
    if content and content_unit:
        content_value = _dec(content)
        normalized_unit = str(content_unit).strip().lower()
        if content_value >= 1000 and normalized_unit in ("g", "gr", "grs", "gs"):
            return f"{quantity} ({format(content_value / 1000, 'f').rstrip('0').rstrip('.')} kg)"
        if content_value >= 1000 and normalized_unit in ("ml", "cc"):
            return f"{quantity} ({format(content_value / 1000, 'f').rstrip('0').rstrip('.')} litros)"
        return f"{quantity} ({content} {content_unit})"
    return quantity


def _snapshot_for_pdf(snapshot: dict) -> dict:
    schools = []
    positive_school_ingredients: set[tuple[object, object]] = set()
    for school in snapshot.get("escuelas", []):
        ingredients = [
            item for item in school.get("ingredientes", [])
            if _dec(item.get("cantidad_final", "0")) > 0
        ]
        positive_school_ingredients.update(
            (school.get("escuela_id"), item.get("ingrediente_id"))
            for item in ingredients
        )
        schools.append({**school, "ingredientes": ingredients})

    providers = []
    for provider in snapshot.get("proveedores", []):
        ingredients = []
        for ingredient in provider.get("ingredientes", []):
            if _dec(ingredient.get("cantidad_total", "0")) <= 0:
                continue
            school_rows = [
                row for row in ingredient.get("escuelas", [])
                if _dec(row.get("cantidad", "0")) > 0
            ]
            if school_rows:
                ingredients.append({**ingredient, "escuelas": school_rows})
        if ingredients:
            providers.append({**provider, "ingredientes": ingredients})

    return {
        **snapshot,
        "escuelas": schools,
        "proveedores": providers,
        "resumen_global": [
            row for row in snapshot.get("resumen_global", [])
            if _dec(row.get("cantidad_total", "0")) > 0
        ],
        "advertencias": [
            warning for warning in snapshot.get("advertencias", [])
            if (warning.get("escuela_id"), warning.get("ingrediente_id"))
            in positive_school_ingredients
        ],
    }


def filtered_snapshot_for_export(
    pedido: GeneracionPedido,
    user: User,
    localidad_id: int | None = None,
    proveedor_id: int | None = None,
    escuela_id: int | None = None,
) -> dict:
    snapshot = school_snapshot_for_user(pedido.datos_snapshot, user)

    selected_school_ids = {
        school.get("escuela_id")
        for school in snapshot.get("escuelas", [])
        if (localidad_id is None or school.get("localidad_id") == localidad_id)
        and (escuela_id is None or school.get("escuela_id") == escuela_id)
    }

    filtered_providers = []
    provider_school_ids: set[int] = set()
    for provider in snapshot.get("proveedores", []):
        if localidad_id is not None and provider.get("localidad_id") != localidad_id:
            continue
        if proveedor_id is not None and provider.get("proveedor_id") != proveedor_id:
            continue

        ingredients = []
        for ingredient in provider.get("ingredientes", []):
            school_rows = [
                row for row in ingredient.get("escuelas", [])
                if row.get("escuela_id") in selected_school_ids
            ]
            if school_rows:
                provider_school_ids.update(row.get("escuela_id") for row in school_rows)
                quantity_total = sum(_dec(row.get("cantidad", "0")) for row in school_rows)
                content_quantity_total = sum(
                    _dec(row.get("cantidad_contenido", "0") or "0") for row in school_rows
                )
                cost_total = quantity_total * _dec(ingredient.get("precio_unitario", "0"))
                ingredients.append({
                    **ingredient,
                    "escuelas": school_rows,
                    "cantidad_total": _qty(quantity_total),
                    "cantidad_contenido_total": (
                        _qty(content_quantity_total)
                        if ingredient.get("contenido_por_unidad") is not None
                        else None
                    ),
                    "costo_total": _money(cost_total),
                })

        if ingredients:
            filtered_providers.append({**provider, "ingredientes": ingredients})

    if proveedor_id is not None:
        selected_school_ids = provider_school_ids

    filtered_schools = [
        school for school in snapshot.get("escuelas", [])
        if school.get("escuela_id") in selected_school_ids
    ]

    resumen_by_key: dict[str, dict] = {}
    total = Decimal("0")
    for school in filtered_schools:
        for item in school.get("ingredientes", []):
            if "proveedor_id" not in item:
                continue
            if localidad_id is not None and item.get("localidad_id") != localidad_id:
                continue
            if proveedor_id is not None and item.get("proveedor_id") != proveedor_id:
                continue

            key = f"{item.get('ingrediente_id')}:{item.get('localidad_id')}:{item.get('proveedor_id')}"
            row = resumen_by_key.setdefault(
                key,
                {
                    "ingrediente_id": item.get("ingrediente_id"),
                    "ingrediente_nombre": item.get("ingrediente_nombre", ""),
                    "unidad": item.get("unidad_final", ""),
                    "localidad_id": item.get("localidad_id"),
                    "localidad_nombre": item.get("localidad_nombre", ""),
                    "proveedor_id": item.get("proveedor_id"),
                    "proveedor_nombre": item.get("proveedor_nombre", ""),
                    "precio_unitario": item.get("precio_unitario", ""),
                    "contenido_por_unidad": item.get("contenido_por_unidad"),
                    "unidad_contenido": item.get("unidad_contenido"),
                    "cantidad_total": Decimal("0"),
                    "cantidad_contenido_total": Decimal("0"),
                    "costo_total": Decimal("0"),
                },
            )
            row["cantidad_total"] += _dec(item.get("cantidad_final", "0"))
            row["cantidad_contenido_total"] += _dec(
                item.get("cantidad_contenido_final", "0") or "0"
            )
            row["costo_total"] += _dec(item.get("costo_total", "0"))
            total += _dec(item.get("costo_total", "0"))

    resumen = [
        {
            **{
                k: v
                for k, v in row.items()
                if k not in ("cantidad_total", "cantidad_contenido_total", "costo_total")
            },
            "cantidad_total": _qty(row["cantidad_total"]),
            "cantidad_contenido_total": (
                _qty(row["cantidad_contenido_total"])
                if row["contenido_por_unidad"] is not None
                else None
            ),
            "costo_total": _money(row["costo_total"]),
        }
        for row in resumen_by_key.values()
    ]

    advertencias = [
        warning for warning in snapshot.get("advertencias", [])
        if warning.get("escuela_id") in selected_school_ids
        and (localidad_id is None or warning.get("localidad_id") == localidad_id)
    ]

    return {
        **snapshot,
        "escuelas": filtered_schools,
        "proveedores": filtered_providers,
        "resumen_global": resumen,
        "advertencias": advertencias,
        "costo_total": _money(total),
    }


def _provider_school_columns(provider: dict) -> list[dict]:
    schools_by_id: dict[int, dict] = {}
    for ingredient in provider.get("ingredientes", []):
        for school in ingredient.get("escuelas", []):
            schools_by_id[school["escuela_id"]] = school
    return sorted(
        schools_by_id.values(),
        key=lambda item: (item.get("escuela_codigo", ""), item.get("escuela_nombre", "")),
    )


def export_resumen_excel(
    pedido: GeneracionPedido,
    user: User,
    localidad_id: int | None = None,
    proveedor_id: int | None = None,
    escuela_id: int | None = None,
) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    snapshot = filtered_snapshot_for_export(pedido, user, localidad_id, proveedor_id, escuela_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen global"
    sheet.append(["RESUMEN SEMANAL SAE"])
    sheet.append(["Semana", snapshot.get("semana_inicio", "")])
    sheet.append(["Dias habiles", ", ".join(str(d) for d in snapshot.get("dias_habiles", []))])
    option = snapshot.get("opcion_menu", {})
    sheet.append(["Menu", f"Opcion {option.get('numero_opcion', '')}"])
    sheet.append(["Costo total", snapshot.get("costo_total", "")])
    sheet.append([])
    sheet.append(["Ingrediente", "Unidad", "Localidad", "Cantidad total", "Proveedor", "Precio unit.", "Costo total"])
    for cell in sheet[7]:
        cell.font = Font(bold=True)

    for row in snapshot.get("resumen_global", []):
        sheet.append([
            row.get("ingrediente_nombre", ""),
            _commercial_unit_label(row),
            row.get("localidad_nombre", ""),
            _commercial_quantity_label(row),
            row.get("proveedor_nombre", ""),
            row.get("precio_unitario", ""),
            row.get("costo_total", ""),
        ])

    if snapshot.get("resumen_global"):
        sheet.append(["TOTAL", "", "", "", "", "", snapshot.get("costo_total", "")])
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

    if snapshot.get("advertencias"):
        sheet.append([])
        sheet.append(["Advertencias"])
        sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)
        sheet.append(["Escuela", "Localidad", "Ingrediente"])
        for warning in snapshot.get("advertencias", []):
            sheet.append([
                warning.get("escuela_nombre", ""),
                warning.get("localidad_nombre", ""),
                warning.get("ingrediente_nombre", ""),
            ])

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 45)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_resumen_pdf(
    pedido: GeneracionPedido,
    user: User,
    localidad_id: int | None = None,
    proveedor_id: int | None = None,
    escuela_id: int | None = None,
) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    snapshot = _snapshot_for_pdf(
        filtered_snapshot_for_export(pedido, user, localidad_id, proveedor_id, escuela_id)
    )
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    option = snapshot.get("opcion_menu", {})
    story = [
        Paragraph("RESUMEN SEMANAL SAE", styles["Title"]),
        Paragraph(f"Semana: {snapshot.get('semana_inicio', '')}", styles["Normal"]),
        Paragraph(f"Menu: Opcion {option.get('numero_opcion', '')}", styles["Normal"]),
        Paragraph(f"Costo total semanal: {snapshot.get('costo_total', '')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [["Ingrediente", "Unidad", "Localidad", "Cant. total", "Proveedor", "Precio unit.", "Costo total"]]
    for row in snapshot.get("resumen_global", []):
        data.append([
            row.get("ingrediente_nombre", ""),
            _commercial_unit_label(row),
            row.get("localidad_nombre", ""),
            _commercial_quantity_label(row),
            row.get("proveedor_nombre", ""),
            row.get("precio_unitario", ""),
            row.get("costo_total", ""),
        ])
    has_rows = len(data) > 1
    if not has_rows:
        data.append(["Sin datos", "", "", "", "", "", ""])
    else:
        data.append(["TOTAL", "", "", "", "", "", snapshot.get("costo_total", "")])

    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    if has_rows:
        table_style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle(table_style))
    story.append(table)

    if snapshot.get("advertencias"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("Ingredientes sin proveedor asignado", styles["Heading2"]))
        warning_data = [["Escuela", "Localidad", "Ingrediente"]]
        for warning in snapshot.get("advertencias", []):
            warning_data.append([
                warning.get("escuela_nombre", ""),
                warning.get("localidad_nombre", ""),
                warning.get("ingrediente_nombre", ""),
            ])
        warning_table = Table(warning_data, repeatRows=1)
        warning_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fef3c7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(warning_table)

    doc.build(story)
    output.seek(0)
    return output


def _provider_excel(snapshot: dict, provider: dict) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    schools = _provider_school_columns(provider)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orden proveedor"
    sheet.append(["ORDEN DE COMPRA"])
    sheet.append(["Semana", snapshot.get("semana_inicio", "")])
    sheet.append(["Proveedor", provider.get("proveedor_nombre", "")])
    sheet.append(["Localidad", provider.get("localidad_nombre", "")])
    sheet.append([])

    header = ["Ingrediente", "Unidad", *[
        f"{school.get('escuela_codigo', '')} - {school.get('escuela_nombre', '')}"
        for school in schools
    ], "TOTAL", "Precio unit.", "Costo estimado"]
    sheet.append(header)
    for cell in sheet[6]:
        cell.font = Font(bold=True)

    ingredientes = provider.get("ingredientes", [])
    for ingredient in ingredientes:
        quantities = {row["escuela_id"]: row.get("cantidad", "") for row in ingredient.get("escuelas", [])}
        sheet.append([
            ingredient.get("ingrediente_nombre", ""),
            _commercial_unit_label(ingredient),
            *[quantities.get(school["escuela_id"], "0.00") for school in schools],
            _commercial_quantity_label(ingredient),
            ingredient.get("precio_unitario", ""),
            ingredient.get("costo_total", ""),
        ])

    if ingredientes:
        total_cost = sum(
            (_dec(item.get("costo_total", "0")) for item in ingredientes),
            Decimal("0"),
        )
        sheet.append([
            "TOTAL",
            "",
            *["" for _ in schools],
            "",
            "",
            _money(total_cost),
        ])
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 35)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _provider_pdf(snapshot: dict, provider: dict) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    schools = _provider_school_columns(provider)
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("ORDEN DE COMPRA", styles["Title"]),
        Paragraph(f"Semana: {snapshot.get('semana_inicio', '')}", styles["Normal"]),
        Paragraph(f"Proveedor: {provider.get('proveedor_nombre', '')}", styles["Normal"]),
        Paragraph(f"Localidad: {provider.get('localidad_nombre', '')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    header = ["Ingrediente", "Unidad", *[
        school.get("escuela_codigo", "") or school.get("escuela_nombre", "")
        for school in schools
    ], "TOTAL"]
    data = [header]
    for ingredient in provider.get("ingredientes", []):
        quantities = {row["escuela_id"]: row.get("cantidad", "") for row in ingredient.get("escuelas", [])}
        data.append([
            ingredient.get("ingrediente_nombre", ""),
            _commercial_unit_label(ingredient),
            *[quantities.get(school["escuela_id"], "0.00") for school in schools],
            _commercial_quantity_label(ingredient),
        ])
    if len(data) == 1:
        data.append(["Sin datos", "", *["" for _ in schools], ""])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)

    total_cost = sum(
        (_dec(item.get("costo_total", "0")) for item in provider.get("ingredientes", [])),
        Decimal("0"),
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"<b>Costo estimado total: {_money(total_cost)}</b>", styles["Normal"]))

    doc.build(story)
    output.seek(0)
    return output


def export_proveedores_zip(
    pedido: GeneracionPedido,
    user: User,
    file_format: str,
    localidad_id: int | None = None,
    proveedor_id: int | None = None,
    escuela_id: int | None = None,
) -> BytesIO:
    snapshot = filtered_snapshot_for_export(pedido, user, localidad_id, proveedor_id, escuela_id)
    if file_format not in ("pdf", "excel"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato invalido")
    if file_format == "pdf":
        snapshot = _snapshot_for_pdf(snapshot)

    output = BytesIO()
    extension = "pdf" if file_format == "pdf" else "xlsx"
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for provider in snapshot.get("proveedores", []):
            base_name = _safe_filename(
                f"{provider.get('proveedor_nombre', 'proveedor')}_{provider.get('localidad_nombre', 'localidad')}"
            )
            content = (
                _provider_pdf(snapshot, provider)
                if file_format == "pdf"
                else _provider_excel(snapshot, provider)
            )
            archive.writestr(f"{base_name}.{extension}", content.getvalue())

    output.seek(0)
    return output


def _providers_by_locality(snapshot: dict) -> list[dict]:
    groups: dict[int, dict] = {}
    for provider in snapshot.get("proveedores", []):
        locality_id = provider.get("localidad_id")
        group = groups.setdefault(
            locality_id,
            {
                "localidad_id": locality_id,
                "localidad_nombre": provider.get("localidad_nombre", ""),
                "proveedores": [],
            },
        )
        group["proveedores"].append(provider)
    return sorted(groups.values(), key=lambda item: item.get("localidad_nombre", ""))


def _locality_pdf(snapshot: dict, locality_group: dict) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("PEDIDO POR LOCALIDAD", styles["Title"]),
        Paragraph(f"Semana: {snapshot.get('semana_inicio', '')}", styles["Normal"]),
        Paragraph(f"Localidad: {locality_group.get('localidad_nombre', '')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [["Proveedor", "Ingrediente", "Unidad", "Escuela", "Cantidad", "Total proveedor"]]
    for provider in locality_group.get("proveedores", []):
        for ingredient in provider.get("ingredientes", []):
            for school in ingredient.get("escuelas", []):
                data.append([
                    provider.get("proveedor_nombre", ""),
                    ingredient.get("ingrediente_nombre", ""),
                    _commercial_unit_label(ingredient),
                    f"{school.get('escuela_codigo', '')} - {school.get('escuela_nombre', '')}",
                    _commercial_quantity_label(
                        {**ingredient, **school},
                        "cantidad",
                        "cantidad_contenido",
                    ),
                    _commercial_quantity_label(ingredient),
                ])
    if len(data) == 1:
        data.append(["Sin datos", "", "", "", "", ""])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)

    total_cost = sum(
        (
            _dec(ingredient.get("costo_total", "0"))
            for provider in locality_group.get("proveedores", [])
            for ingredient in provider.get("ingredientes", [])
        ),
        Decimal("0"),
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"<b>Costo estimado total: {_money(total_cost)}</b>", styles["Normal"]))

    doc.build(story)
    output.seek(0)
    return output


def export_localidades_pdf_zip(
    pedido: GeneracionPedido,
    user: User,
    localidad_id: int | None = None,
    proveedor_id: int | None = None,
    escuela_id: int | None = None,
) -> BytesIO:
    snapshot = _snapshot_for_pdf(
        filtered_snapshot_for_export(pedido, user, localidad_id, proveedor_id, escuela_id)
    )
    output = BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for locality_group in _providers_by_locality(snapshot):
            base_name = _safe_filename(locality_group.get("localidad_nombre", "localidad"))
            archive.writestr(f"{base_name}.pdf", _locality_pdf(snapshot, locality_group).getvalue())
    output.seek(0)
    return output


def _school_pdf(snapshot: dict, school: dict) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("PEDIDO POR ESCUELA", styles["Title"]),
        Paragraph(f"Semana: {snapshot.get('semana_inicio', '')}", styles["Normal"]),
        Paragraph(
            f"Escuela: {school.get('codigo', '')} - {school.get('nombre', '')}",
            styles["Normal"],
        ),
        Paragraph(f"Localidad: {school.get('localidad_nombre', '')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [["Ingrediente", "Unidad", "Cantidad", "Proveedor", "Precio unit.", "Costo"]]
    school_total = Decimal("0")
    for item in school.get("ingredientes", []):
        if "proveedor_id" not in item:
            continue
        school_total += _dec(item.get("costo_total", "0"))
        data.append([
            item.get("ingrediente_nombre", ""),
            _commercial_unit_label(item),
            _commercial_quantity_label(
                item,
                "cantidad_final",
                "cantidad_contenido_final",
            ),
            item.get("proveedor_nombre", ""),
            item.get("precio_unitario", ""),
            item.get("costo_total", ""),
        ])
    has_rows = len(data) > 1
    if not has_rows:
        data.append(["Sin datos", "", "", "", "", ""])
    else:
        data.append(["TOTAL", "", "", "", "", _money(school_total)])

    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    if has_rows:
        table_style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle(table_style))
    story.append(table)
    doc.build(story)
    output.seek(0)
    return output


def export_escuelas_pdf_zip(
    pedido: GeneracionPedido,
    user: User,
    localidad_id: int | None = None,
    proveedor_id: int | None = None,
    escuela_id: int | None = None,
) -> BytesIO:
    snapshot = _snapshot_for_pdf(
        filtered_snapshot_for_export(pedido, user, localidad_id, proveedor_id, escuela_id)
    )
    output = BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for school in snapshot.get("escuelas", []):
            base_name = _safe_filename(
                f"{school.get('codigo', '')}_{school.get('nombre', 'escuela')}"
            )
            archive.writestr(f"{base_name}.pdf", _school_pdf(snapshot, school).getvalue())
    output.seek(0)
    return output


def export_pedido_excel(pedido: GeneracionPedido, user: User) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    snapshot = school_snapshot_for_user(pedido.datos_snapshot, user)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Semana", snapshot.get("semana_inicio", "")])
    summary.append(["Dias habiles", ", ".join(str(d) for d in snapshot.get("dias_habiles", []))])
    option = snapshot.get("opcion_menu", {})
    summary.append(["Menu", f"Opcion {option.get('numero_opcion', '')}"])
    summary.append(["Costo total", snapshot.get("costo_total", "")])
    summary.append([])
    summary.append(["Ingrediente", "Unidad", "Localidad", "Proveedor", "Precio unit.", "Cantidad", "Costo"])
    for cell in summary[6]:
        cell.font = Font(bold=True)
    for row in snapshot.get("resumen_global", []):
        summary.append([
            row.get("ingrediente_nombre", ""),
            _commercial_unit_label(row),
            row.get("localidad_nombre", ""),
            row.get("proveedor_nombre", ""),
            row.get("precio_unitario", ""),
            _commercial_quantity_label(row),
            row.get("costo_total", ""),
        ])

    if snapshot.get("resumen_global"):
        summary.append(["TOTAL", "", "", "", "", "", snapshot.get("costo_total", "")])
        for cell in summary[summary.max_row]:
            cell.font = Font(bold=True)

    providers_sheet = workbook.create_sheet("Proveedores")
    providers_sheet.append(["Proveedor", "Localidad", "Ingrediente", "Unidad", "Escuela", "Cantidad", "Total proveedor"])
    for cell in providers_sheet[1]:
        cell.font = Font(bold=True)
    for provider in snapshot.get("proveedores", []):
        for ingredient in provider.get("ingredientes", []):
            for school in ingredient.get("escuelas", []):
                providers_sheet.append([
                    provider.get("proveedor_nombre", ""),
                    provider.get("localidad_nombre", ""),
                    ingredient.get("ingrediente_nombre", ""),
                    _commercial_unit_label(ingredient),
                    f"{school.get('escuela_codigo', '')} - {school.get('escuela_nombre', '')}",
                    _commercial_quantity_label(
                        {**ingredient, **school},
                        "cantidad",
                        "cantidad_contenido",
                    ),
                    _commercial_quantity_label(ingredient),
                ])

    warnings_sheet = workbook.create_sheet("Advertencias")
    warnings_sheet.append(["Tipo", "Escuela", "Localidad", "Ingrediente"])
    for cell in warnings_sheet[1]:
        cell.font = Font(bold=True)
    for warning in snapshot.get("advertencias", []):
        warnings_sheet.append([
            warning.get("tipo", ""),
            warning.get("escuela_nombre", ""),
            warning.get("localidad_nombre", ""),
            warning.get("ingrediente_nombre", ""),
        ])

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 45)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_pedido_pdf(pedido: GeneracionPedido, user: User) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    snapshot = _snapshot_for_pdf(school_snapshot_for_user(pedido.datos_snapshot, user))
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Resumen semanal SAE", styles["Title"]),
        Paragraph(f"Semana: {snapshot.get('semana_inicio', '')}", styles["Normal"]),
        Paragraph(f"Costo total: {snapshot.get('costo_total', '')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    summary_data = [["Ingrediente", "Unidad", "Localidad", "Proveedor", "Cantidad", "Costo"]]
    for row in snapshot.get("resumen_global", []):
        summary_data.append([
            row.get("ingrediente_nombre", ""),
            _commercial_unit_label(row),
            row.get("localidad_nombre", ""),
            row.get("proveedor_nombre", ""),
            _commercial_quantity_label(row),
            row.get("costo_total", ""),
        ])
    has_summary_rows = len(summary_data) > 1
    if not has_summary_rows:
        summary_data.append(["Sin datos", "", "", "", "", ""])
    else:
        summary_data.append(["TOTAL", "", "", "", "", snapshot.get("costo_total", "")])

    summary_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    if has_summary_rows:
        summary_style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
    summary_table = Table(summary_data, repeatRows=1)
    summary_table.setStyle(TableStyle(summary_style))
    story.append(summary_table)

    for provider in snapshot.get("proveedores", []):
        story.append(Spacer(1, 14))
        story.append(Paragraph(
            f"{provider.get('proveedor_nombre', '')} - {provider.get('localidad_nombre', '')}",
            styles["Heading2"],
        ))
        provider_data = [["Ingrediente", "Unidad", "Escuela", "Cantidad"]]
        for ingredient in provider.get("ingredientes", []):
            for school in ingredient.get("escuelas", []):
                provider_data.append([
                    ingredient.get("ingrediente_nombre", ""),
                    _commercial_unit_label(ingredient),
                    f"{school.get('escuela_codigo', '')} - {school.get('escuela_nombre', '')}",
                    _commercial_quantity_label(
                        {**ingredient, **school},
                        "cantidad",
                        "cantidad_contenido",
                    ),
                ])
        table = Table(provider_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(table)

    if snapshot.get("advertencias"):
        story.append(Spacer(1, 14))
        story.append(Paragraph("Advertencias", styles["Heading2"]))
        warning_data = [["Escuela", "Localidad", "Ingrediente"]]
        for warning in snapshot.get("advertencias", []):
            warning_data.append([
                warning.get("escuela_nombre", ""),
                warning.get("localidad_nombre", ""),
                warning.get("ingrediente_nombre", ""),
            ])
        warning_table = Table(warning_data, repeatRows=1)
        warning_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fef3c7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(warning_table)

    doc.build(story)
    output.seek(0)
    return output
